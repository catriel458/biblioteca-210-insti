import json
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from collections import Counter
from .models import Libro, Inventario, Mapas, Multimedia, Notebook, Proyector, Varios, Prestamo, Sancion
from .forms import LibroForm, MapaForm, MultimediaForm, NotebookForm, ProyectorForm, VariosForm
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import csv
import io  # Agregar esta línea
from django.contrib import messages #Para mensajes
from django.http import JsonResponse
from django.db.models import Q  # Añade esta línea
from django.contrib import messages
from django.utils import timezone
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import RegistroForm, LoginForm, CambiarPasswordForm
from .models import Usuario


import csv
import io
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import Libro, Mapas, Multimedia, Notebook, Proyector, Varios

import csv
import io
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import Libro, Mapas, Multimedia, Notebook, Proyector, Varios

from django.contrib.auth.models import User
from django.contrib.auth import get_user_model


import threading
import time
import datetime
from django.conf import settings

from io import StringIO
from django.core.management import call_command
from django.views.decorators.csrf import csrf_exempt


from django.db import transaction

import logging

# Configurar logger
logger = logging.getLogger(__name__)

# Variables globales para control de thread
ultima_verificacion = None
thread_verificacion = None

def verificacion_automatica():
    """Ejecuta verificación cada 2 minutos en background"""
    global ultima_verificacion
    
    print("[VERIFICACION] Iniciando sistema automático cada 2 minutos")
    
    while True:
        ahora = datetime.datetime.now()

        
        # Verificar si han pasado 2 minutos (120 segundos)
        if (ultima_verificacion is None or 
            (ahora - ultima_verificacion).total_seconds() >= 120):
            
            try:
                print(f"[{ahora.strftime('%H:%M:%S')}] Ejecutando verificación automática...")
                verificar_y_notificar_vencimientos()
                ultima_verificacion = ahora
                print(f"[{ahora.strftime('%H:%M:%S')}] Verificación completada")
            except Exception as e:
                print(f"[{ahora.strftime('%H:%M:%S')}] ERROR en verificación: {e}")
        
        # Dormir 30 segundos antes de verificar nuevamente
        time.sleep(30)

def calcular_fecha_devolucion_exacta(fecha_inicio, dias=15):
    """
    Calcula fecha de devolución exacta preservando hora, minutos y segundos
    """
    fecha_devolucion = fecha_inicio + datetime.timedelta(
        days=dias,
        hours=0,
        minutes=0,
        seconds=0
    )
    
    print(f"[DEBUG TIEMPO] Fecha inicio: {fecha_inicio.strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"[DEBUG TIEMPO] Fecha devolución: {fecha_devolucion.strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"[DEBUG TIEMPO] Diferencia exacta: {(fecha_devolucion - fecha_inicio).total_seconds()} segundos")
    
    return fecha_devolucion

def es_bibliotecaria(user):
    """Verifica si el usuario es bibliotecaria"""
    return user.is_authenticated and user.perfil == 'bibliotecaria'

# AGREGAR este decorador personalizado (después de la función es_bibliotecaria existente)
def bibliotecaria_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not hasattr(request.user, 'es_bibliotecaria') or not request.user.es_bibliotecaria:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('lista_libros')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def es_usuario_activo(user):
    """Verifica si el usuario está autenticado y activo"""
    return user.is_authenticated and user.is_active




def success_view(request):
    return render(request, 'libros/success.html')

# Busqueda libros

@login_required(login_url='login')
def buscar_libros(request):
    query = request.GET.get('q', '')
    libros = Libro.objects.filter(
        Q(titulo__icontains=query) | 
        Q(autor__icontains=query) | 
        Q(resumen__icontains=query),
        estado='Disponible'
    ).values(
        'id_libro',  # IMPORTANTE: Asegúrate de que este campo esté incluido
        # 'num_inventario',
        'titulo', 
        'autor', 
        'editorial', 
        'clasificacion_cdu', 
        'siglas_autor_titulo', 
        'sede', 
        'disponibilidad', 
        'img'
    )

    # Para depuración, puedes imprimir los datos
    print("Datos de libros:", list(libros))
    
    return JsonResponse(list(libros), safe=False)

# Buscar de mapas
def buscar_mapas(request):
    query = request.GET.get('q', '')
    mapas = Mapas.objects.filter(
        Q(tipo__icontains=query) | 
        Q(descripcion__icontains=query),
        estado='Disponible'
    ).values('id_mapa', 'tipo', 'descripcion', 'num_ejemplar')

    return JsonResponse(list(mapas), safe=False)

# Buscador de multimedia
def buscar_multimedia(request):
    query = request.GET.get('q', '')
    multimedia = Multimedia.objects.filter(
        Q(materia__icontains=query) | 
        Q(contenido__icontains=query),
        estado='Disponible'
    ).values('id_multi', 'materia', 'contenido', 'num_ejemplar')

    return JsonResponse(list(multimedia), safe=False) 

# Buscador de notebooks
# Código viejo
'''def buscar_notebooks(request):
    query = request.GET.get('q', '')
    notebooks = Notebook.objects.filter(
        Q(marca__icontains=query) | 
        Q(modelo__icontains=query)
    ).values('id_netbook', 'marca', 'modelo', 'num_ejemplar')

    return JsonResponse(list(notebooks), safe=False)'''

# Código modificado (funcionando)
def buscar_notebooks(request):
    query = request.GET.get('q', '')
    if query:
        notebooks = Notebook.objects.filter(marca_not__icontains=query, estado='Disponible') | Notebook.objects.filter(modelo_not__icontains=query, estado='Disponible')
    else:
        notebooks = Notebook.objects.filter(estado='Disponible')

    data = list(notebooks.values('id_not', 'marca_not', 'modelo_not', 'num_ejemplar'))
    return JsonResponse(data, safe=False)

# Buscador de proyectores

def buscar_proyectores(request):
    query = request.GET.get('q', '')
    proyectores = Proyector.objects.filter(
        Q(marca_pro__icontains=query) | Q(modelo_pro__icontains=query),
        estado='Disponible'  # Asegúrate de que esto coincide con el campo en tu modelo
    ).values('id_proyector', 'marca_pro', 'modelo_pro', 'num_ejemplar')

    return JsonResponse(list(proyectores), safe=False)

# Buscador de varios
def buscar_varios(request):
    query = request.GET.get('q', '')
    varios = Varios.objects.filter(
        Q(tipo__icontains=query),
        estado='Disponible' 
    ).values('id_varios', 'tipo', 'num_ejemplar')

    return JsonResponse(list(varios), safe=False)

# Borrar libros
@user_passes_test(es_bibliotecaria, login_url='login')
def borrar_libros(request):
    if request.method == 'POST':
        Libro.objects.all().delete()
        messages.success(request, "Todos los libros han sido borrados.")
        return redirect('lista_libros')  # Cambia esto a la URL donde quieras redirigir después de borrar

    return render(request, 'libros/borrar_libros.html')  # Crea un template para confirmar la acción

# Métodos de biblioteca:


def alta_inventario(request):
    pass


def baja_inventario(request):
    pass


def modificacion_inventario(request):
    pass


def asignar_prestamo(request):
    pass


def cancelar_prestamo(request):
    pass


def modificar_prestamo(request):
    pass

# Método alumno y profesor:


# Pantalla principal

# Vista para la pantalla principal:


@login_required
def pantalla_principal(request):
    """Vista de la pantalla principal con información contextual"""
    global thread_verificacion
    
    # Iniciar thread de verificación si no está corriendo
    if thread_verificacion is None or not thread_verificacion.is_alive():
        thread_verificacion = threading.Thread(target=verificacion_automatica, daemon=True)
        thread_verificacion.start()
        print("[SISTEMA] Thread de verificación automática iniciado")
    
    context = {}
    
    # Si es bibliotecaria, agregar estadísticas de sanciones
    if request.user.es_bibliotecaria():
        from .models import Sancion
        context['sanciones_pendientes_count'] = Sancion.objects.filter(estado='pendiente').count()
        context['sanciones_confirmadas_count'] = Sancion.objects.filter(estado='confirmada').count()
        
        # Verificar préstamos vencidos para actualizar sanciones
        verificar_prestamos_vencidos()
    
    return render(request, 'libros/pantalla_principal.html', context)

# Libros


# Vista para dar de alta un libro:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_libro(request):
    form = LibroForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        libro = form.save(commit=False)
        libro.save()
        context = {'form': form, 'success': 'Libro registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_libro.html', context)

# Vista para dar de baja un libro:

@user_passes_test(es_bibliotecaria, login_url='login')
def baja_libro(request):
    if request.method == 'POST':
        libro_id = request.POST.get('libro_id')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del libro
        # Cambiado para asegurar que sea Libro
        libro = get_object_or_404(Libro, id_libro=libro_id)
        libro.estado = 'No disponible'  # Asegúrate de cambiar el estado
        libro.motivo_baja = motivo_baja
          # ✅ NUEVA LÓGICA: Guardar imagen en base de datos
        if imagen_rota:
            libro.imagen_rota_nombre = imagen_rota.name
            libro.imagen_rota_contenido = imagen_rota.read()  # Lee los bytes
            libro.imagen_rota_tipo = imagen_rota.content_type  # Ej: 'image/jpeg'
        
        libro.save()

        # Redirigir a la lista de libros después de la baja
        return redirect('lista_libros')

    return redirect('lista_libros')

@login_required
def ver_imagen_baja(request, libro_id):
    """Vista para mostrar la imagen de baja almacenada en la base de datos"""
    libro = get_object_or_404(Libro, id_libro=libro_id)
    
    if not libro.imagen_rota_contenido:
        return HttpResponse("No hay imagen disponible", status=404)
    
    # Retornar la imagen como respuesta HTTP
    response = HttpResponse(libro.imagen_rota_contenido, content_type=libro.imagen_rota_tipo or 'image/jpeg')
    response['Content-Disposition'] = f'inline; filename="{libro.imagen_rota_nombre or "imagen.jpg"}"'
    return response

# Vista para editar un libro:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_libro(request, libro_id):
    libro = get_object_or_404(Libro, id_libro=libro_id)

    if request.method == 'POST':
        form = LibroForm(request.POST, instance=libro)
        if form.is_valid():
            form.save()
            return redirect('lista_libros')
    else:
        form = LibroForm(instance=libro)

    return render(request, 'libros/editar_libro.html', {'form': form, 'libro': libro})

# Carga masiva libros:

@login_required
@csrf_exempt  # Solo para esta vista específica
def cargar_csv_lote(request):
    """Procesa un lote de registros (máximo 50 por request)"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            registros = data.get('registros', [])
            
            if not registros:
                return JsonResponse({'success': False, 'error': 'No hay registros'})
            
            creados = 0
            errores = []
            
            with transaction.atomic():
                for idx, row in enumerate(registros, start=1):
                    try:
                        tipo_material = row.get('tipo_material', '').strip().lower()
                        
                        if tipo_material == 'libro':
                            titulo = row.get('titulo', '').strip()
                            if not titulo:
                                errores.append(f"Registro {idx}: Falta título")
                                continue
                            
                            def safe_int(value, default=1):
                                if value is None or value == '':
                                    return default
                                try:
                                    return int(str(value).strip())
                                except (ValueError, TypeError):
                                    return default
                            
                            libro = Libro.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=safe_int(row.get('num_ejemplar'), 1),
                                titulo=titulo,
                                autor=row.get('autor', 'Desconocido').strip() or 'Desconocido',
                                editorial=row.get('editorial', 'Sin editorial').strip() or 'Sin editorial',
                                clasificacion_cdu=row.get('clasificacion_cdu', 'Sin clasificar').strip() or 'Sin clasificar',
                                siglas_autor_titulo=row.get('siglas_autor_titulo', 'ABC').strip() or 'ABC',
                                num_inventario=safe_int(row.get('num_inventario'), 1),
                                resumen=row.get('resumen', 'Sin resumen').strip() or 'Sin resumen',
                                etiqueta_palabra_clave=row.get('etiqueta_palabra_clave', '').strip() or '',
                                sede=row.get('sede', 'La Plata').strip() or 'La Plata',
                                disponibilidad='Disponible',
                                observaciones=row.get('observaciones', '').strip() or '',
                                img=row.get('img', '').strip() or None
                            )
                            creados += 1
                        
                        # Agregar otros tipos (mapa, multimedia, etc.) igual que antes
                        
                    except Exception as e:
                        errores.append(f"Registro {idx}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'creados': creados,
                'errores': errores
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required(login_url='login')
def cargar_csv(request):
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        
        if not csv_file:
            messages.error(request, "No se seleccionó ningún archivo.")
            return redirect('lista_libros')
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "El archivo debe ser CSV.")
            return redirect('lista_libros')

        try:
            # Leer archivo con diferentes encodings
            try:
                decoded_file = csv_file.read().decode('utf-8')
            except UnicodeDecodeError:
                csv_file.seek(0)
                decoded_file = csv_file.read().decode('latin-1')
            
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            rows = list(reader)
            
            if not rows:
                messages.error(request, "El archivo CSV está vacío.")
                return redirect('lista_libros')

            print(f"\n{'='*60}")
            print(f"Procesando {len(rows)} filas del CSV")
            print(f"{'='*60}\n")

            libros_creados = 0
            errores = []

            # Función auxiliar para convertir a entero de forma segura
            def safe_int(value, default=1):
                """Convierte un valor a entero, retorna default si está vacío o es inválido"""
                if value is None or value == '':
                    return default
                try:
                    return int(str(value).strip())
                except (ValueError, TypeError):
                    return default

            with transaction.atomic():
                for idx, row in enumerate(rows, start=1):
                    try:
                        tipo_material = row.get('tipo_material', '').strip().lower()
                        
                        if tipo_material == 'libro':
                            titulo = row.get('titulo', '').strip()
                            
                            if not titulo:
                                errores.append(f"Fila {idx}: Falta título")
                                continue
                            
                            # Usar safe_int para campos numéricos
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            num_inventario = safe_int(row.get('num_inventario'), 1)
                            
                            libro = Libro.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                titulo=titulo,
                                autor=row.get('autor', 'Desconocido').strip() or 'Desconocido',
                                editorial=row.get('editorial', 'Sin editorial').strip() or 'Sin editorial',
                                clasificacion_cdu=row.get('clasificacion_cdu', 'Sin clasificar').strip() or 'Sin clasificar',
                                siglas_autor_titulo=row.get('siglas_autor_titulo', 'ABC').strip() or 'ABC',
                                num_inventario=num_inventario,
                                resumen=row.get('resumen', 'Sin resumen').strip() or 'Sin resumen',
                                etiqueta_palabra_clave=row.get('etiqueta_palabra_clave', '').strip() or '',
                                sede=row.get('sede', 'La Plata').strip() or 'La Plata',
                                disponibilidad='Disponible',
                                observaciones=row.get('observaciones', '').strip() or '',
                                img=row.get('img', '').strip() or None
                            )
                            
                            libros_creados += 1
                            if libros_creados % 10 == 0:  # Log cada 10 libros
                                print(f"✅ {libros_creados} libros creados...")
                        
                        elif tipo_material == 'mapa':
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            
                            mapa = Mapas.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                tipo=row.get('tipo', 'Sin tipo').strip() or 'Sin tipo'
                            )
                            libros_creados += 1
                        
                        elif tipo_material == 'multimedia':
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            
                            multimedia = Multimedia.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                materia=row.get('materia', 'Sin materia').strip() or 'Sin materia',
                                contenido=row.get('contenido', 'Sin contenido').strip() or 'Sin contenido'
                            )
                            libros_creados += 1
                        
                        elif tipo_material == 'notebook':
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            
                            notebook = Notebook.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                marca_not=row.get('marca_not', 'Sin marca').strip() or 'Sin marca',
                                modelo_not=row.get('modelo_not', 'Sin modelo').strip() or 'Sin modelo'
                            )
                            libros_creados += 1
                        
                        elif tipo_material == 'proyector':
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            
                            proyector = Proyector.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                marca_pro=row.get('marca_pro', 'Sin marca').strip() or 'Sin marca',
                                modelo_pro=row.get('modelo_pro', 'Sin modelo').strip() or 'Sin modelo'
                            )
                            libros_creados += 1
                        
                        elif tipo_material == 'varios':
                            num_ejemplar = safe_int(row.get('num_ejemplar'), 1)
                            
                            varios = Varios.objects.create(
                                estado='Disponible',
                                motivo_baja=row.get('motivo_baja', '').strip() or None,
                                descripcion=row.get('descripcion', '').strip() or None,
                                num_ejemplar=num_ejemplar,
                                tipo=row.get('tipo', 'Sin tipo').strip() or 'Sin tipo'
                            )
                            libros_creados += 1
                        
                        else:
                            if tipo_material:  # Solo reportar si hay un valor
                                errores.append(f"Fila {idx}: Tipo '{tipo_material}' no reconocido")
                    
                    except Exception as e:
                        error_msg = f"Fila {idx}: {str(e)}"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")

            print(f"\n{'='*60}")
            print(f"✅ COMPLETADO: {libros_creados} registros creados")
            print(f"❌ Errores: {len(errores)}")
            print(f"{'='*60}\n")

            # Mensajes al usuario
            if libros_creados > 0:
                messages.success(request, f"✅ {libros_creados} registros cargados exitosamente.")
            
            if errores and len(errores) <= 10:
                # Si hay pocos errores, mostrarlos todos
                for error in errores:
                    messages.warning(request, error)
            elif errores:
                # Si hay muchos errores, mostrar resumen
                messages.warning(request, f"⚠️ Se encontraron {len(errores)} errores. Primeros 3:")
                for error in errores[:3]:
                    messages.error(request, error)
            
            if libros_creados == 0 and len(errores) == 0:
                messages.error(request, "❌ No se reconoció ningún registro válido en el CSV.")
            
            return redirect('lista_libros')
        
        except Exception as e:
            print(f"\n❌ ERROR CRÍTICO: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return redirect('lista_libros')

    return render(request, 'libros/upload_csv.html')

# Mapas

# Vista para listar mapas disponibles:


def mapas_view(request):
    mapas = Mapas.objects.filter(estado='Disponible')
    return render(request, 'libros/mapas.html', {'mapas': mapas})

# Vista para dar de baja un mapa:


def baja_mapa(request):
    if request.method == 'POST':
        mapa_id = request.POST.get('mapa_id')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del mapa
        mapa = get_object_or_404(Mapas, id_mapa=mapa_id)
        mapa.estado = 'No disponible'
        mapa.motivo_baja = motivo_baja
        if imagen_rota:
            mapa.imagen_rota = imagen_rota
        mapa.save()

        return redirect('mapas')

    return redirect('mapas')

# Vista para dar de alta un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_mapa(request):
    form = MapaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        mapa = form.save(commit=False)
        mapa.save()
        context = {'form': form, 'success': 'Mapa registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_mapa.html', context)

# Vista para editar un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_mapa(request, mapa_id):
    mapa = get_object_or_404(Mapas, id_mapa=mapa_id)

    if request.method == 'POST':
        form = MapaForm(request.POST, instance=mapa)
        if form.is_valid():
            form.save()
            return redirect('mapas')
    else:
        form = MapaForm(instance=mapa)

    return render(request, 'libros/editar_mapa.html', {'form': form, 'mapa': mapa})

# Vista para mostrar elementos multimedia (por implementar):

# Multimedia


def multimedia_view(request):
    multimedia = Multimedia.objects.filter(estado='Disponible')
    return render(request, 'libros/multimedia.html', {'multimedia': multimedia})

# Vista para dar de baja un mapa:


def baja_multimedia(request):
    if request.method == 'POST':
        multi_id = request.POST.get('id_multi')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del mapa
        multimedia = get_object_or_404(Multimedia, id_multi=multi_id)
        multimedia.estado = 'No disponible'
        multimedia.motivo_baja = motivo_baja
        if imagen_rota:
            multimedia.imagen_rota = imagen_rota
        multimedia.save()

        return redirect('multimedia')

    return redirect('multimedia')

# Vista para dar de alta un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_multimedia(request):
    form = MultimediaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        multimedia = form.save(commit=False)
        multimedia.save()
        context = {'form': form,
                   'success': 'Multimedia registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_multimedia.html', context)

# Vista para editar un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_multimedia(request, multi_id):
    multimedia = get_object_or_404(Multimedia, id_multi=multi_id)

    if request.method == 'POST':
        form = MultimediaForm(request.POST, instance=multimedia)
        if form.is_valid():
            form.save()
            return redirect('multimedia')
    else:
        form = MultimediaForm(instance=multimedia)

    return render(request, 'libros/editar_multimedia.html', {'form': form, 'mapa': multimedia})

# Notebook


def notebook_view(request):
    notebook = Notebook.objects.filter(estado='Disponible')
    return render(request, 'libros/notebook.html', {'notebook': notebook})

# Vista para dar de baja un mapa:


def baja_notebook(request):
    if request.method == 'POST':
        not_id = request.POST.get('not_id')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del mapa
        notebook = get_object_or_404(Notebook, id_not=not_id)
        notebook.estado = 'No disponible'
        notebook.motivo_baja = motivo_baja
        if imagen_rota:
            notebook.imagen_rota = imagen_rota
        notebook.save()

        return redirect('notebook')

    return redirect('notebook')

# Vista para dar de alta un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_notebook(request):
    form = NotebookForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        notebook = form.save(commit=False)
        notebook.save()
        context = {'form': form,
                   'success': 'Notebook registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_notebook.html', context)

# Vista para editar un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_notebook(request, not_id):
    notebook = get_object_or_404(Notebook, id_not=not_id)

    if request.method == 'POST':
        form = NotebookForm(request.POST, instance=notebook)
        if form.is_valid():
            form.save()
            return redirect('notebook')
    else:
        form = NotebookForm(instance=notebook)

    return render(request, 'libros/editar_notebook.html', {'form': form, 'notebook': notebook})

# Proyector


def proyector_view(request):
    proyector = Proyector.objects.filter(estado='Disponible')
    return render(request, 'libros/proyector.html', {'proyector': proyector})

# Vista para dar de baja un mapa:


def baja_proyector(request):
    if request.method == 'POST':
        proyector_id = request.POST.get('proyector_id')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del mapa
        proyector = get_object_or_404(Proyector, id_proyector=proyector_id)
        proyector.estado = 'No disponible'
        proyector.motivo_baja = motivo_baja
        if imagen_rota:
            proyector.imagen_rota = imagen_rota
        proyector.save()

        return redirect('proyector')

    return redirect('proyector')

# Vista para dar de alta un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_proyector(request):
    form = ProyectorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        proyector = form.save(commit=False)
        proyector.save()
        context = {'form': form,
                   'success': 'Proyector registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_proyector.html', context)

# Vista para editar un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_proyector(request, proyector_id):
    proyector = get_object_or_404(Proyector, id_proyector=proyector_id)

    if request.method == 'POST':
        form = ProyectorForm(request.POST, instance=proyector)
        if form.is_valid():
            form.save()
            return redirect('proyector')
    else:
        form = ProyectorForm(instance=proyector)

    return render(request, 'libros/editar_proyector.html', {'form': form, 'proyector': proyector})


# Varios


def varios_view(request):
    varios = Varios.objects.filter(estado='Disponible')
    return render(request, 'libros/varios.html', {'varios': varios})

# Vista para dar de baja un mapa:


def baja_varios(request):
    if request.method == 'POST':
        varios_id = request.POST.get('varios_id')
        motivo_baja = request.POST.get('motivo_baja')
        imagen_rota = request.FILES.get('imagen_rota')

        # Lógica para actualizar el estado del mapa
        varios = get_object_or_404(Varios, id_varios=varios_id)
        varios.estado = 'No disponible'
        varios.motivo_baja = motivo_baja
        if imagen_rota:
            varios.imagen_rota = imagen_rota
        varios.save()

        return redirect('varios')

    return redirect('varios')

# Vista para dar de alta un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def alta_varios(request):
    form = VariosForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        varios = form.save(commit=False)
        varios.save()
        context = {'form': form,
                   'success': 'Varios registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}

    return render(request, 'libros/alta_varios.html', context)

# Vista para editar un mapa:

@user_passes_test(es_bibliotecaria, login_url='login')
def editar_varios(request, varios_id):
    varios = get_object_or_404(Varios, id_varios=varios_id)

    if request.method == 'POST':
        form = VariosForm(request.POST, instance=varios)
        if form.is_valid():
            form.save()
            return redirect('varios')
    else:
        form = VariosForm(instance=varios)

    return render(request, 'libros/editar_varios.html', {'form': form, 'varios': varios})


# Vista para registro de bajas 

def registro_bajas(request):
    # Filtrar solo los libros que están 'No disponible'
    libros_no_disponibles = Libro.objects.filter(estado='No disponible')

    # Contexto a pasar al template
    context = {
        'libros_no_disponibles': libros_no_disponibles,
    }

    return render(request, 'libros/registro_bajas.html', context)

# Reactivar libros en el historial de bajas
def reactivar_libro(request, libro_id):
    if request.method == 'POST':
        libro = get_object_or_404(Libro, id_libro=libro_id)
        libro.estado = 'Disponible'
        libro.motivo_baja = ''  # Opcional: limpiar el motivo de baja
        libro.save()
        messages.success(request, f'El libro "{libro.titulo}" ha sido reactivado exitosamente.')
        return redirect('registro_de_bajas')
    return redirect('registro_de_bajas')

# PRESTAMOS

# def solicitar_prestamo(request, libro_id):
#     libro = get_object_or_404(Libro, id_libro=libro_id)
    
#     # Verificar si el libro está disponible
#     if libro.estado != 'Disponible':
#         messages.error(request, "Este libro no está disponible para préstamo.")
#         return redirect('lista_libros')
    
#     # Crear el préstamo
#     if request.method == 'POST':
#         nombre_usuario = request.POST.get('nombre_usuario', '')
#         email_usuario = request.POST.get('email_usuario', '')
#         tipo_usuario = request.POST.get('tipo_usuario', 'alumno')
#         tipo_prestamo = request.POST.get('tipo_prestamo', 'domicilio')
        
#         # NO calcular fecha límite aquí - se calculará cuando se apruebe
#         prestamo = Prestamo(
#             nombre_usuario=nombre_usuario,
#             email_usuario=email_usuario,
#             libro=libro,
#             tipo_prestamo=tipo_prestamo,
#             tipo_usuario=tipo_usuario,
#             estado='solicitado',
#             fecha_limite_reserva=None  # Se establecerá cuando se apruebe la solicitud
#         )
#         prestamo.save()
        
#         # Cambiar estado del libro a reservado
#         libro.estado = 'Reservado'
#         libro.save()
        
#         messages.success(request, f"Has solicitado el préstamo del libro '{libro.titulo}'. La bibliotecaria revisará tu solicitud y tendrás 3 días hábiles para retirarlo una vez aprobada.")
#         return redirect('prestamos_solicitados')
    
#     return render(request, 'libros/solicitar_prestamo.html', {'libro': libro})


def aprobar_solicitud_prestamo(request, prestamo_id):
    """
    Aprueba la solicitud de préstamo y empieza el cronómetro de reserva
    """
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado != 'solicitado':
        messages.error(request, f"La solicitud no puede ser aprobada porque su estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    # AQUÍ es donde empieza el cronómetro de reserva (3 días hábiles)
    fecha_limite = calcular_dias_habiles(timezone.now(), 3)
    
    prestamo.estado = 'aprobado_reserva'  # Nuevo estado: aprobado para reserva
    prestamo.fecha_limite_reserva = fecha_limite
    prestamo.fecha_aprobacion = timezone.now()
    prestamo.save()
    
    # Calcular días hábiles restantes para mostrar en el mensaje
    dias_habiles = 3
    messages.success(request, f"✅ RESERVA APROBADA: '{prestamo.libro.titulo}' para {prestamo.nombre_usuario}. Tiempo límite: {fecha_limite.strftime('%d/%m/%Y a las %H:%M')} ({dias_habiles} días hábiles).")
    
    return redirect('gestionar_prestamos')

def cancelar_reserva_usuario(request, prestamo_id):
    """
    Permite al usuario cancelar su propia solicitud o reserva
    """
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    # Solo permitir cancelar si está en estado solicitado o aprobado_reserva
    if prestamo.estado not in ['solicitado', 'aprobado_reserva']:
        messages.error(request, f"No puedes cancelar este préstamo porque su estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', 'Cancelado por el usuario')
        
        # Cambiar estado a rechazado
        prestamo.estado = 'rechazado'
        prestamo.motivo_rechazo = f"CANCELADO POR USUARIO: {motivo_cancelacion}"
        prestamo.save()
        
        # Devolver el libro al estado disponible
        libro = prestamo.libro
        libro.estado = 'Disponible'
        libro.save()
        
        messages.success(request, f"Has cancelado exitosamente la reserva del libro '{prestamo.libro.titulo}'.")
        
        return redirect('gestionar_prestamos')
    
    return render(request, 'libros/cancelar_reserva_usuario.html', {'prestamo': prestamo})

@login_required
def prestamos_solicitados(request):
    # Verificar préstamos vencidos antes de mostrar
    verificar_prestamos_vencidos()
    
    # Filtrar préstamos según el tipo de usuario
    if request.user.es_bibliotecaria():
        # La bibliotecaria ve todos los préstamos
        prestamos = Prestamo.objects.all().order_by('-fecha_solicitud')
    else:
        # Los usuarios normales solo ven sus propios préstamos
        prestamos = Prestamo.objects.filter(
            Q(email_usuario=request.user.email) | Q(usuario=request.user)
        ).order_by('-fecha_solicitud')
    
    # AGREGAR INFORMACIÓN DE DEBUG para cada préstamo
    for prestamo in prestamos:
        if prestamo.usuario and prestamo.usuario.perfil == 'docente':
            print(f"[DEBUG] Préstamo {prestamo.id_prestamo}:")
            print(f"  - Usuario: {prestamo.usuario.get_full_name()} (perfil: {prestamo.usuario.perfil})")
            print(f"  - Estado: {prestamo.estado}")
            print(f"  - Extensión ya solicitada: {prestamo.extension_solicitada}")
            print(f"  - Puede extender: {prestamo.puede_extender_prestamo()}")
            
            if prestamo.fecha_devolucion_programada:
                from django.utils import timezone
                tiempo_restante = prestamo.fecha_devolucion_programada - timezone.now()
                dias_restantes = tiempo_restante.days
                print(f"  - Días restantes: {dias_restantes}")
    
    # Verificar si hay alertas de tiempo (menos de 1 día hábil restante)
    alertas = []
    ahora = timezone.now()
    
    for prestamo in prestamos:
        # Alerta para reservas próximas a vencer
        if prestamo.estado == 'aprobado_reserva' and prestamo.fecha_limite_reserva:
            tiempo_restante = prestamo.fecha_limite_reserva - ahora
            if tiempo_restante.total_seconds() > 0 and tiempo_restante.total_seconds() < 86400:
                alertas.append(f"¡ATENCIÓN! La reserva del libro '{prestamo.libro.titulo}' vence pronto.")
        
        # NUEVA ALERTA: Préstamos de docentes que pueden extender
        if (prestamo.usuario == request.user and 
            prestamo.usuario.perfil == 'docente' and 
            prestamo.estado == 'aprobado' and 
            prestamo.puede_extender_prestamo()):
            alertas.append(f"💡 Puedes EXTENDER tu préstamo del libro '{prestamo.libro.titulo}' por 15 días más.")
    
    return render(request, 'libros/prestamos_solicitados.html', {
        'prestamos': prestamos,
        'alertas': alertas,
        'user': request.user  # Asegurar que el usuario esté disponible en el template
    })

# MODIFICAR ESTA FUNCIÓN EXISTENTE
def rechazar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado not in ['solicitado', 'vencido']:
        messages.error(request, f"El préstamo no puede ser rechazado porque su estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    if request.method == 'POST':
        motivo_rechazo = request.POST.get('motivo_rechazo', '')
        prestamo.estado = 'rechazado'
        prestamo.motivo_rechazo = motivo_rechazo
        prestamo.save()
        
        # Devolver el libro al estado disponible
        libro = prestamo.libro
        libro.estado = 'Disponible'
        libro.save()
        
        messages.success(request, f"El préstamo del libro '{prestamo.libro.titulo}' ha sido rechazado.")
        
        return redirect('gestionar_prestamos')
    
    return render(request, 'libros/rechazar_prestamo.html', {'prestamo': prestamo})

@user_passes_test(es_bibliotecaria)
def finalizar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado != 'aprobado':
        messages.error(request, f"El préstamo no puede ser finalizado porque su estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    prestamo.estado = 'finalizado'
    prestamo.fecha_devolucion_real = timezone.now()
    prestamo.save()
    
    # Devolver el libro al estado disponible
    libro = prestamo.libro
    libro.estado = 'Disponible'
    libro.save()
    
    messages.success(request, f"El préstamo del libro '{prestamo.libro.titulo}' ha sido finalizado.")
    
    return redirect('gestionar_prestamos')


# FUNCIÓN PARA CALCULAR DÍAS HÁBILES (Lunes a Viernes)
def calcular_dias_habiles(fecha_inicio, dias_habiles):
    """
    Calcula una fecha que esté N días hábiles después de la fecha de inicio
    """
    dias_agregados = 0
    fecha_actual = fecha_inicio
    
    while dias_agregados < dias_habiles:
        fecha_actual += datetime.timedelta(days=1)
        # Si es día de semana (lunes=0, domingo=6), contar como día hábil
        if fecha_actual.weekday() < 5:  # 0-4 son lunes a viernes
            dias_agregados += 1
    
    return fecha_actual

# Modificar la función verificar_prestamos_vencidos

def verificar_prestamos_vencidos():
    """
    Verifica y actualiza préstamos que han vencido
    """
    ahora = timezone.now()
    
    print(f"[DEBUG VENCIMIENTOS] Verificando a las: {ahora.strftime('%d/%m/%Y %H:%M:%S')}")
    
    # Buscar préstamos activos que han vencido
    prestamos_activos_vencidos = Prestamo.objects.filter(
        estado='aprobado',
        fecha_devolucion_programada__lt=ahora  # Comparación exacta
    )
    
    print(f"[DEBUG VENCIMIENTOS] Préstamos vencidos encontrados: {prestamos_activos_vencidos.count()}")
    
    for prestamo in prestamos_activos_vencidos:
        print(f"[DEBUG VENCIMIENTOS] Procesando préstamo {prestamo.id_prestamo}:")
        print(f"  - Usuario: {prestamo.nombre_usuario}")
        print(f"  - Libro: {prestamo.libro.titulo}")
        print(f"  - Fecha programada: {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y %H:%M:%S')}")
        print(f"  - Ahora: {ahora.strftime('%d/%m/%Y %H:%M:%S')}")
        
        diferencia = ahora - prestamo.fecha_devolucion_programada
        print(f"  - Vencido por: {diferencia.total_seconds()} segundos")
        
        # Cambiar estado a vencido
        prestamo.estado = 'vencido'
        prestamo.save()
        
        # Crear sanción si tiene usuario
        if prestamo.usuario:
            prestamo.crear_sancion_por_vencimiento()
            print(f"  - Sanción creada para {prestamo.usuario.get_full_name()}")
        else:
            print(f"  - No se puede crear sanción: sin usuario asignado")
    
    # También verificar préstamos de reserva vencidos
    prestamos_reserva_vencidos = Prestamo.objects.filter(
        estado='aprobado_reserva',
        fecha_limite_reserva__lt=ahora
    )
    
    for prestamo in prestamos_reserva_vencidos:
        prestamo.estado = 'vencido'
        prestamo.save()
        
        # Devolver libro a disponible
        prestamo.libro.estado = 'Disponible'
        prestamo.libro.save()
        
        print(f"[DEBUG VENCIMIENTOS] Reserva vencida: {prestamo.id_prestamo}")

# Nueva función para verificar vencimientos en tiempo real
def verificar_vencimientos_tiempo_real():
    """
    Función optimizada que solo verifica préstamos que pueden estar vencidos
    """
    ahora = timezone.now()
    
    # Solo verificar préstamos que tienen fechas de vencimiento próximas (últimas 2 horas)
    hace_2_horas = ahora - datetime.timedelta(hours=2)
    
    # Préstamos activos que pueden haber vencido recientemente
    prestamos_candidatos = Prestamo.objects.filter(
        estado='aprobado',
        fecha_devolucion_programada__gte=hace_2_horas,
        fecha_devolucion_programada__lt=ahora
    )
    
    if prestamos_candidatos.exists():
        print(f"[TIEMPO_REAL] Verificando {prestamos_candidatos.count()} préstamos candidatos a vencimiento")
        verificar_prestamos_vencidos()
    else:
        print(f"[TIEMPO_REAL] No hay préstamos candidatos a vencimiento en las últimas 2 horas")

# Mejorar la función que establece la fecha de devolución para mayor precisión
def establecer_fecha_devolucion_precisa(prestamo, dias=15):
    """
    Establece fecha de devolución con hora específica (ej: 18:00)
    """
    fecha_base = timezone.now()
    
    # Opción 1: Exactamente X días desde ahora (preserva hora exacta)
    fecha_devolucion = fecha_base + datetime.timedelta(days=dias)
    
    # Opción 2: X días a las 18:00 (hora fija) - descomenta si prefieres esto
    # fecha_devolucion = fecha_base + datetime.timedelta(days=dias)
    # fecha_devolucion = fecha_devolucion.replace(hour=18, minute=0, second=0, microsecond=0)
    
    prestamo.fecha_devolucion_programada = fecha_devolucion
    prestamo.save()
    
    print(f"[DEBUG] Fecha de devolución establecida para préstamo {prestamo.id_prestamo}: {fecha_devolucion}")
    return fecha_devolucion


# AGREGAR ESTAS FUNCIONES NUEVAS

def confirmar_retiro_libro(request, prestamo_id):
    """
    Confirma que el alumno retiró el libro físicamente
    """
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado not in ['aprobado_reserva']:
        messages.error(request, f"No se puede confirmar el retiro porque el estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    if prestamo.fecha_limite_reserva and timezone.now() > prestamo.fecha_limite_reserva:
        messages.error(request, "El tiempo de reserva ha vencido. No se puede confirmar el retiro.")
        return redirect('gestionar_prestamos')
    
    if request.method == 'POST':
        ahora_retiro = timezone.now()
        
        libro = prestamo.libro
        libro.estado = 'No disponible'
        libro.save()
        prestamo.estado = 'aprobado'
        prestamo.fecha_retiro_real = ahora_retiro
        
        # ✅ LÍNEA CORREGIDA
        prestamo.fecha_devolucion_programada = calcular_fecha_devolucion_exacta(ahora_retiro, 15)
        
        prestamo.fecha_limite_reserva = None
        prestamo.save()
        
        # Mensaje con hora exacta
        fecha_format = prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y a las %H:%M:%S')
        messages.success(request, f"Se confirmó el retiro del libro '{prestamo.libro.titulo}'. Fecha de devolución: {fecha_format}.")
        
        return redirect('gestionar_prestamos')
    
    return render(request, 'libros/confirmar_retiro.html', {'prestamo': prestamo})

def marcar_no_retiro(request, prestamo_id):
    """
    Marca que el alumno no retiró el libro
    """
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado != 'solicitado':
        messages.error(request, f"No se puede marcar como no retirado porque el estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', 'No retirado por el usuario')
        
        # Cambiar estado a rechazado
        prestamo.estado = 'rechazado'
        prestamo.motivo_rechazo = motivo
        prestamo.save()
        
        # Devolver el libro al estado disponible
        libro = prestamo.libro
        libro.estado = 'Disponible'
        libro.save()
        
        messages.success(request, f"Se marcó como no retirado el libro '{prestamo.libro.titulo}'. El libro está nuevamente disponible.")
        
        return redirect('gestionar_prestamos')
    
    return render(request, 'libros/marcar_no_retiro.html', {'prestamo': prestamo})

# AGREGAR al archivo libros/views.py

@user_passes_test(es_bibliotecaria)
def marcar_libro_devuelto(request, sancion_id):
    """
    Marca que el usuario devolvió el libro y finaliza la sanción
    """
    sancion = get_object_or_404(Sancion, id_sancion=sancion_id)
    
    if sancion.estado != 'confirmada':
        messages.error(request, f"No se puede marcar como devuelto porque el estado actual es {sancion.get_estado_display()}.")
        return redirect('gestionar_sanciones')
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        
        # Marcar sanción como cumplida
        sancion.estado = 'cumplida'
        sancion.fecha_finalizacion = timezone.now()
        sancion.observaciones_bibliotecaria = f"Libro devuelto. {observaciones}".strip()
        sancion.save()
        
        # Finalizar el préstamo relacionado
        prestamo = sancion.prestamo
        if prestamo.estado in ['vencido', 'aprobado']:
            prestamo.estado = 'finalizado'
            prestamo.fecha_devolucion_real = timezone.now()
            prestamo.save()
        
        # Marcar libro como disponible
        libro = prestamo.libro
        libro.estado = 'Disponible'
        libro.save()
        
        # Enviar email de notificación al usuario (opcional)
        try:
            send_mail(
                subject='Sanción finalizada - Biblioteca ISFD 210',
                message=f'''
Estimado/a {sancion.usuario.get_full_name()},

Te informamos que tu sanción ha sido finalizada exitosamente.

📚 Libro: {sancion.prestamo.libro.titulo}
✅ Estado: Sanción cumplida
📅 Fecha de resolución: {timezone.now().strftime('%d/%m/%Y %H:%M')}

Ya puedes inscribirte normalmente a las mesas de final.

¡Gracias por resolver tu situación!

Saludos,
Biblioteca ISFD 210
                ''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[sancion.usuario.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error enviando email de finalización: {e}")
        
        messages.success(request, f"✅ Libro devuelto confirmado. Sanción finalizada para {sancion.usuario.get_full_name()}. El libro '{libro.titulo}' está nuevamente disponible.")
        return redirect('gestionar_sanciones')
    
    # Si es GET, mostrar página de confirmación
    context = {
        'sancion': sancion,
        'libro': sancion.prestamo.libro,
        'usuario': sancion.usuario,
    }
    return render(request, 'libros/marcar_libro_devuelto.html', context)

# MODIFICAR la vista gestionar_sanciones para incluir contadores adicionales
@user_passes_test(es_bibliotecaria)
def gestionar_sanciones(request):
    """Vista principal para gestionar sanciones con filtros mejorados"""
    verificar_vencimientos_tiempo_real()
    
    filtro = request.GET.get('filtro', 'pendientes')
    filtro_tipo = request.GET.get('filtro_tipo', 'todos')  # NUEVO FILTRO
    
    # Filtro base por estado
    if filtro == 'pendientes':
        sanciones = Sancion.objects.filter(estado='pendiente')
    elif filtro == 'confirmadas':
        sanciones = Sancion.objects.filter(estado='confirmada')
    elif filtro == 'canceladas':
        sanciones = Sancion.objects.filter(estado='cancelada')
    elif filtro == 'cumplidas':
        sanciones = Sancion.objects.filter(estado='cumplida')
    else:
        sanciones = Sancion.objects.all()
    
    # Filtro adicional por tipo de usuario de sanción
    if filtro_tipo == 'alumno':
        sanciones = sanciones.filter(tipo_usuario_sancion='alumno')
    elif filtro_tipo == 'docente':
        sanciones = sanciones.filter(tipo_usuario_sancion='docente')
    
    sanciones = sanciones.order_by('-fecha_creacion')
    
    # Estadísticas
    total_pendientes = Sancion.objects.filter(estado='pendiente').count()
    total_confirmadas = Sancion.objects.filter(estado='confirmada').count()
    total_canceladas = Sancion.objects.filter(estado='cancelada').count()
    total_cumplidas = Sancion.objects.filter(estado='cumplida').count()
    
    # Nuevas estadísticas por tipo
    total_alumnos = Sancion.objects.filter(tipo_usuario_sancion='alumno').count()
    total_docentes = Sancion.objects.filter(tipo_usuario_sancion='docente').count()
    docentes_suspendidos = Sancion.objects.filter(
        tipo_usuario_sancion='docente',
        estado='confirmada',
        fecha_suspension_hasta__gt=timezone.now()
    ).count()
    
    context = {
        'sanciones': sanciones,
        'filtro': filtro,
        'filtro_tipo': filtro_tipo,  # NUEVO CONTEXTO
        'total_pendientes': total_pendientes,
        'total_confirmadas': total_confirmadas,
        'total_canceladas': total_canceladas,
        'total_cumplidas': total_cumplidas,
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'docentes_suspendidos': docentes_suspendidos,
    }
    return render(request, 'libros/gestionar_sanciones.html', context)

# AGREGAR al archivo libros/urls.py
# En la sección de URLs de sanciones:

    # URLs de sanciones
    path('gestionar-sanciones/', views.gestionar_sanciones, name='gestionar_sanciones'),
    path('confirmar-sancion/<int:sancion_id>/', views.confirmar_sancion, name='confirmar_sancion'),
    path('cancelar-sancion/<int:sancion_id>/', views.cancelar_sancion, name='cancelar_sancion'),
    path('marcar-libro-devuelto/<int:sancion_id>/', views.marcar_libro_devuelto, name='marcar_libro_devuelto'),  # NUEVA URL
    path('mis-sanciones/', views.mis_sanciones, name='mis_sanciones'),

# Agregar estos imports al inicio de tu archivo views.py
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import RegistroForm, LoginForm, CambiarPasswordForm
from .models import Usuario

# Funciones auxiliares para verificar permisos
def es_bibliotecaria(user):
    return user.is_authenticated and user.perfil == 'bibliotecaria'

# Agregar estas vistas al final de tu archivo views.py

def login_view(request):
    if request.user.is_authenticated:
        return redirect('pantalla_principal')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            dni = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=dni, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bienvenido/a {user.get_full_name()}')
                return redirect('pantalla_principal')
    else:
        form = LoginForm()
    
    return render(request, 'libros/login.html', {'form': form})

# En views.py, modifica registro_view temporalmente
def registro_view(request):
    # TEMPORAL: Hacer que el primer usuario sea bibliotecaria
    primer_usuario = Usuario.objects.count() == 0
    
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            if primer_usuario:
                user.perfil = 'bibliotecaria'
                user.is_staff = True
                user.is_superuser = True
            user.save()
            messages.success(request, 'Cuenta creada exitosamente.')
            return redirect('login')
    else:
        form = RegistroForm()
    
    return render(request, 'libros/registro.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')

@login_required
def perfil_usuario(request):
    return render(request, 'libros/perfil_usuario.html', {'usuario': request.user})

@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = CambiarPasswordForm(request.user, request.POST)
        if form.is_valid():
            request.user.set_password(form.cleaned_data['password_nueva'])
            request.user.save()
            messages.success(request, 'Contraseña cambiada exitosamente. Inicia sesión nuevamente.')
            return redirect('login')
    else:
        form = CambiarPasswordForm(request.user)
    
    return render(request, 'libros/cambiar_password.html', {'form': form})

# Actualizar estas vistas existentes para agregar autenticación


@login_required  # Agregar este decorador
def lista_libros(request):
    # Obtener todos los libros disponibles
    todos_los_libros = Libro.objects.filter(estado='Disponible').extra(
    select={'titulo_clean': "TRIM(UPPER(titulo))"}
    ).order_by('titulo_clean')
    
    # Configurar paginación con 30 libros por página
    paginator = Paginator(todos_los_libros, 30)
    
    # Obtener el número de página desde GET
    page = request.GET.get('page')
    
    try:
        libros = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, muestra la primera página
        libros = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera de rango, muestra la última página
        libros = paginator.page(paginator.num_pages)
    
    context = {
        'libros': libros,
        'paginator': paginator,
        'page_obj': libros,  # Para compatibilidad con templates
    }
    
    return render(request, 'libros/lista_libros.html', context)


@user_passes_test(es_bibliotecaria)  # Solo bibliotecarias pueden dar de alta libros
def alta_libro(request):
    # ... mantener el código existente ...
    form = LibroForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        libro = form.save(commit=False)
        libro.save()
        context = {'form': form, 'success': 'Libro registrado exitosamente.'}
    else:
        context = {'form': form, 'error': 'Por favor complete todos los campos obligatorios.'} if request.method == 'POST' else {
            'form': form}
    return render(request, 'libros/alta_libro.html', context)


# Actualizar la vista de préstamos solicitados para mostrar solo los del usuario actual
@login_required
def prestamos_solicitados(request):
    # Verificar préstamos vencidos antes de mostrar
    verificar_prestamos_vencidos()
    
    # Filtrar préstamos según el tipo de usuario
    if request.user.es_bibliotecaria():
        # La bibliotecaria ve todos los préstamos
        prestamos = Prestamo.objects.all().order_by('-fecha_solicitud')
    else:
        # Los usuarios normales solo ven sus propios préstamos
        prestamos = Prestamo.objects.filter(
            Q(email_usuario=request.user.email) | Q(usuario=request.user)
        ).order_by('-fecha_solicitud')
    
    # Verificar si hay alertas de tiempo (menos de 1 día hábil restante)
    alertas = []
    ahora = timezone.now()
    
    for prestamo in prestamos:
        if prestamo.estado == 'solicitado' and prestamo.fecha_limite_reserva:
            tiempo_restante = prestamo.fecha_limite_reserva - ahora
            # Si queda menos de 24 horas
            if tiempo_restante.total_seconds() > 0 and tiempo_restante.total_seconds() < 86400:  # 24 horas en segundos
                alertas.append(f"¡ATENCIÓN! El préstamo del libro '{prestamo.libro.titulo}' vence pronto.")
    
    return render(request, 'libros/prestamos_solicitados.html', {
        'prestamos': prestamos,
        'alertas': alertas
    })

# Actualizar la vista de solicitar préstamo para usar el usuario logueado

@login_required
def solicitar_prestamo(request, libro_id):
    libro = get_object_or_404(Libro, id_libro=libro_id)
    
    # Verificar si el libro está disponible
    if libro.estado != 'Disponible':
        messages.error(request, "Este libro no está disponible para préstamo.")
        return redirect('lista_libros')
    
    # NUEVA VERIFICACIÓN: Límite de préstamos simultáneos
    if not request.user.puede_solicitar_mas_prestamos():
        messages.error(request, 
            "Has alcanzado el límite máximo de 3 préstamos simultáneos. "
            "Debes devolver un libro antes de solicitar otro.")
        return redirect('mis_prestamos_activos')
    
    # NUEVA VERIFICACIÓN: Si el usuario puede solicitar préstamos
    if not request.user.puede_solicitar_prestamo():
        messages.error(request, "No puedes solicitar préstamos debido a sanciones activas.")
        return redirect('mis_sanciones')
    
    # NUEVA VERIFICACIÓN: Si tiene sanciones críticas
    if request.user.tiene_sanciones_criticas():
        messages.warning(request, "Tienes sanciones pendientes de compensación.")
        return redirect('verificar_sanciones_criticas')
    
    # Crear el préstamo
    if request.method == 'POST':
        tipo_usuario = request.POST.get('tipo_usuario', 'alumno')
        tipo_prestamo = request.POST.get('tipo_prestamo', 'domicilio')
        
        prestamo = Prestamo(
            usuario=request.user,
            nombre_usuario=request.user.get_full_name(),
            email_usuario=request.user.email,
            libro=libro,
            tipo_prestamo=tipo_prestamo,
            tipo_usuario=tipo_usuario,
            estado='solicitado',
            fecha_limite_reserva=None
        )
        prestamo.save()
        
        prestamos_restantes = request.user.get_prestamos_disponibles() - 1  # -1 porque acabamos de solicitar uno
        
        mensaje_base = f"Has solicitado el préstamo del libro '{libro.titulo}'. La biblioteca revisará tu solicitud."
        if prestamos_restantes > 0:
            mensaje_base += f" Puedes solicitar {prestamos_restantes} préstamos más."
        else:
            mensaje_base += " Has alcanzado el límite de 3 préstamos simultáneos."
        
        messages.success(request, mensaje_base)
        return redirect('prestamos_solicitados')
    
    # Pasar información adicional al template
    prestamos_activos = request.user.get_prestamos_activos().count()
    prestamos_disponibles = request.user.get_prestamos_disponibles()
    
    context = {
        'libro': libro,
        'prestamos_activos': prestamos_activos,
        'prestamos_disponibles': prestamos_disponibles
    }
    
    return render(request, 'libros/solicitar_prestamo.html', context)

# Gestión de préstamos solo para bibliotecarias
@user_passes_test(es_bibliotecaria)
def gestionar_prestamos(request):
    # Verificar préstamos vencidos antes de mostrar (ahora con precisión)
    verificar_vencimientos_tiempo_real()
    
    # Obtener préstamos según filtro
    filtro = request.GET.get('filtro', 'todos')
    
    if filtro == 'solicitados':
        prestamos = Prestamo.objects.filter(estado='solicitado').order_by('-fecha_solicitud')
    elif filtro == 'activos':
        prestamos = Prestamo.objects.filter(estado='aprobado').order_by('-fecha_aprobacion')
    elif filtro == 'finalizados':
        prestamos = Prestamo.objects.filter(estado__in=['finalizado', 'rechazado', 'vencido']).order_by('-fecha_solicitud')
    else:
        prestamos = Prestamo.objects.all().order_by('-fecha_solicitud')
    
    return render(request, 'libros/gestionar_prestamos.html', {
        'prestamos': prestamos,
        'filtro': filtro
    })

# Agregar el decorador a todas las funciones de gestión de préstamos
@user_passes_test(es_bibliotecaria)
def aprobar_prestamo(request, prestamo_id):
    # ... código existente ...
    pass

@user_passes_test(es_bibliotecaria)
def rechazar_prestamo(request, prestamo_id):
    # ... código existente ...
    pass



# AGREGAR estas nuevas vistas al final del archivo

@bibliotecaria_required
def gestion_usuarios(request):
    """Vista principal para gestión de usuarios"""
    usuarios = Usuario.objects.all().order_by('fecha_registro')
    context = {
        'usuarios': usuarios,
        'total_usuarios': usuarios.count(),
        'bibliotecarias': usuarios.filter(perfil='bibliotecaria').count(),
        'alumnos': usuarios.filter(perfil='alumno').count(),
    }
    return render(request, 'libros/gestion_usuarios.html', context)

@bibliotecaria_required
def buscar_usuarios(request):
    """Vista AJAX para búsqueda de usuarios"""
    query = request.GET.get('q', '')
    usuarios = Usuario.objects.filter(
        Q(dni__icontains=query) |
        Q(nombre__icontains=query) |
        Q(apellido__icontains=query) |
        Q(email__icontains=query)
    ).values(
        'id',
        'dni',
        'nombre',
        'apellido', 
        'email',
        'perfil',
        'is_active',
        'fecha_registro'
    )
    
    # Convertir fecha_registro a string para JSON
    usuarios_list = []
    for usuario in usuarios:
        usuario['fecha_registro'] = usuario['fecha_registro'].strftime('%d/%m/%Y')
        usuarios_list.append(usuario)
    
    return JsonResponse(usuarios_list, safe=False)

@bibliotecaria_required
def crear_usuario(request):
    """Vista para crear nuevo usuario"""
    if request.method == 'POST':
        try:
            dni = request.POST.get('dni')
            nombre = request.POST.get('nombre')
            apellido = request.POST.get('apellido')
            email = request.POST.get('email')
            password = request.POST.get('password')
            perfil = request.POST.get('perfil', 'alumno')
            
            # Validaciones
            if Usuario.objects.filter(dni=dni).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'El DNI ya existe'
                })
            
            if Usuario.objects.filter(email=email).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'El email ya está registrado'
                })
            
            # Crear usuario
            usuario = Usuario.objects.create_user(
                dni=dni,
                password=password,
                nombre=nombre,
                apellido=apellido,
                email=email,
                perfil=perfil
            )
            
            return JsonResponse({
                'success': True, 
                'message': 'Usuario creado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error al crear usuario: {str(e)}'
            })
    
    return render(request, 'libros/crear_usuario.html')

@bibliotecaria_required
def editar_usuario(request, usuario_id):
    """Vista para editar usuario existente"""
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        try:
            usuario.nombre = request.POST.get('nombre')
            usuario.apellido = request.POST.get('apellido')
            usuario.email = request.POST.get('email')
            usuario.perfil = request.POST.get('perfil', 'alumno')
            usuario.is_active = request.POST.get('is_active') == 'on'
            
            # Cambiar contraseña solo si se proporciona
            new_password = request.POST.get('password')
            if new_password and new_password.strip():
                usuario.set_password(new_password)
            
            usuario.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Usuario actualizado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error al actualizar usuario: {str(e)}'
            })
    
    context = {'usuario': usuario}
    return render(request, 'libros/editar_usuario.html', context)

@bibliotecaria_required
def eliminar_usuario(request):
    """Vista para eliminar usuario"""
    if request.method == 'POST':
        try:
            usuario_id = request.POST.get('usuario_id')
            usuario = get_object_or_404(Usuario, id=usuario_id)
            
            # No permitir eliminar la última bibliotecaria
            if usuario.perfil == 'bibliotecaria':
                bibliotecarias_count = Usuario.objects.filter(perfil='bibliotecaria').count()
                if bibliotecarias_count <= 1:
                    return JsonResponse({
                        'success': False,
                        'message': 'No se puede eliminar la única bibliotecaria del sistema'
                    })
            
            # No permitir que se elimine a sí mismo
            if usuario.id == request.user.id:
                return JsonResponse({
                    'success': False,
                    'message': 'No puedes eliminarte a ti mismo'
                })
            
            dni = usuario.dni
            usuario.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Usuario {dni} eliminado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar usuario: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@bibliotecaria_required
def exportar_usuarios_excel(request):
    """Vista para exportar usuarios a Excel"""
    try:
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios del Sistema"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'ID', 'DNI', 'Nombre', 'Apellido', 'Email', 
            'Tipo', 'Estado', 'Fecha Registro', 'Último Acceso'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos de usuarios
        usuarios = Usuario.objects.all().order_by('fecha_registro')
        for row, usuario in enumerate(usuarios, 2):
            ws.cell(row=row, column=1, value=usuario.id)
            ws.cell(row=row, column=2, value=usuario.dni)
            ws.cell(row=row, column=3, value=usuario.nombre)
            ws.cell(row=row, column=4, value=usuario.apellido)
            ws.cell(row=row, column=5, value=usuario.email)
            ws.cell(row=row, column=6, value="Bibliotecaria" if usuario.perfil == 'bibliotecaria' else "Alumno")
            ws.cell(row=row, column=7, value="Activo" if usuario.is_active else "Inactivo")
            ws.cell(row=row, column=8, value=usuario.fecha_registro.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=9, value=usuario.last_login.strftime('%d/%m/%Y %H:%M') if usuario.last_login else 'Nunca')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=usuarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar usuarios: {str(e)}')
        return redirect('gestion_usuarios')
    
# Nueva función para enviar avisos de vencimiento
def enviar_avisos_vencimiento():
    """
    Envía emails de aviso un día antes del vencimiento
    """
    mañana = timezone.now() + datetime.timedelta(days=1)
    inicio_mañana = mañana.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_mañana = mañana.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Préstamos que vencen mañana
    prestamos_por_vencer = Prestamo.objects.filter(
        estado='aprobado',
        fecha_devolucion_programada__range=(inicio_mañana, fin_mañana)
    )
    
    for prestamo in prestamos_por_vencer:
        try:
            send_mail(
                subject='Recordatorio: Devolución de libro mañana',
                message=f'''
Hola {prestamo.nombre_usuario},

Te recordamos que mañana {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y')} 
vence el plazo para devolver el libro:

📚 "{prestamo.libro.titulo}"
✍️ Autor: {prestamo.libro.autor}

Por favor, acércate a la biblioteca para realizar la devolución y evitar sanciones.

Saludos,
Biblioteca ISFD 210
                ''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[prestamo.email_usuario],
                fail_silently=False,
            )
        except Exception as e:
            print(f"Error enviando email a {prestamo.email_usuario}: {e}")

# Nuevas vistas para gestión de sanciones
@user_passes_test(es_bibliotecaria)
@user_passes_test(es_bibliotecaria)
def gestionar_sanciones(request):
    """Vista principal para gestionar sanciones con filtros mejorados"""
    verificar_vencimientos_tiempo_real()
    
    filtro = request.GET.get('filtro', 'pendientes')
    filtro_tipo = request.GET.get('filtro_tipo', 'todos')  # ASEGURAR VALOR POR DEFECTO
    
    # Filtro base por estado
    if filtro == 'pendientes':
        sanciones = Sancion.objects.filter(estado='pendiente')
    elif filtro == 'confirmadas':
        sanciones = Sancion.objects.filter(estado='confirmada')
    elif filtro == 'canceladas':
        sanciones = Sancion.objects.filter(estado='cancelada')
    elif filtro == 'cumplidas':
        sanciones = Sancion.objects.filter(estado='cumplida')
    else:
        sanciones = Sancion.objects.all()
    
    # Filtro adicional por tipo de usuario de sanción
    if filtro_tipo == 'alumno':
        sanciones = sanciones.filter(tipo_usuario_sancion='alumno')
    elif filtro_tipo == 'docente':
        sanciones = sanciones.filter(tipo_usuario_sancion='docente')
    # Si es 'todos', no filtrar por tipo
    
    sanciones = sanciones.order_by('-fecha_creacion')
    
    # Estadísticas
    total_pendientes = Sancion.objects.filter(estado='pendiente').count()
    total_confirmadas = Sancion.objects.filter(estado='confirmada').count()
    total_canceladas = Sancion.objects.filter(estado='cancelada').count()
    total_cumplidas = Sancion.objects.filter(estado='cumplida').count()
    
    # Nuevas estadísticas por tipo
    total_alumnos = Sancion.objects.filter(tipo_usuario_sancion='alumno').count()
    total_docentes = Sancion.objects.filter(tipo_usuario_sancion='docente').count()
    docentes_suspendidos = Sancion.objects.filter(
        tipo_usuario_sancion='docente',
        estado='confirmada',
        fecha_suspension_hasta__gt=timezone.now()
    ).count()
    
    context = {
        'sanciones': sanciones,
        'filtro': filtro,
        'filtro_tipo': filtro_tipo,  # ASEGURAR QUE SE PASE AL TEMPLATE
        'total_pendientes': total_pendientes,
        'total_confirmadas': total_confirmadas,
        'total_canceladas': total_canceladas,
        'total_cumplidas': total_cumplidas,
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'docentes_suspendidos': docentes_suspendidos,
    }
    return render(request, 'libros/gestionar_sanciones.html', context)

@user_passes_test(es_bibliotecaria)
def confirmar_sancion(request, sancion_id):
    """Confirma una sanción pendiente con tipo de usuario"""
    sancion = get_object_or_404(Sancion, id_sancion=sancion_id)
    
    if sancion.estado != 'pendiente':
        messages.error(request, f"La sanción no puede ser confirmada porque su estado actual es {sancion.get_estado_display()}.")
        return redirect('gestionar_sanciones')
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        tipo_usuario_sancion = request.POST.get('tipo_usuario_sancion', 'alumno')
        
        sancion.estado = 'confirmada'
        sancion.fecha_confirmacion = timezone.now()
        sancion.observaciones_bibliotecaria = observaciones
        sancion.tipo_usuario_sancion = tipo_usuario_sancion
        
        # Lógica específica para docentes
        if tipo_usuario_sancion == 'docente':
            # Calcular fecha de suspensión (3 meses)
            sancion.fecha_suspension_hasta = sancion.calcular_fecha_suspension_docente()
            
            # Actualizar acumulado de sanciones para este usuario
            sanciones_docente = Sancion.objects.filter(
                usuario=sancion.usuario,
                tipo_usuario_sancion='docente',
                estado='confirmada'
            ).count() + 1  # +1 porque esta se está confirmando ahora
            
            sancion.acumulado_sanciones = sanciones_docente
            
            # Verificar si llegó a 3 sanciones
            if sanciones_docente >= 3:
                messages.warning(
                    request, 
                    f"⚠️ ALERTA: {sancion.usuario.get_full_name()} ha alcanzado {sanciones_docente} sanciones. "
                    f"Debe devolver el material y comprar un material extra como compensación."
                )
        
        sancion.save()
        
        # Enviar email de notificación
        try:
            if tipo_usuario_sancion == 'docente':
                fecha_suspension = sancion.fecha_suspension_hasta.strftime('%d/%m/%Y') if sancion.fecha_suspension_hasta else 'N/A'
                mensaje_email = f'''
Estimado/a {sancion.usuario.get_full_name()},

Se ha aplicado una SANCIÓN DOCENTE a tu cuenta:

📚 Libro: {sancion.prestamo.libro.titulo}
📅 Motivo: {sancion.motivo}
⚠️ Sanción: Suspensión de préstamos por 3 meses
📆 Suspendido hasta: {fecha_suspension}
📊 Sanciones acumuladas: {sancion.acumulado_sanciones}

Esta sanción te inhabilitará para solicitar préstamos hasta la fecha indicada.
{"¡IMPORTANTE! Has alcanzado 3 sanciones. Debes devolver el material y comprar un material extra como compensación." if sancion.acumulado_sanciones >= 3 else ""}

Para resolver esta situación, acércate a la biblioteca.

Saludos,
Biblioteca ISFD 210
                '''
            else:
                mensaje_email = f'''
Estimado/a {sancion.usuario.get_full_name()},

Se ha aplicado una sanción a tu cuenta:

📚 Libro: {sancion.prestamo.libro.titulo}
📅 Motivo: {sancion.motivo}
⚠️ Sanción: {sancion.get_tipo_sancion_display()}

Esta sanción te inhabilitará para inscribirte a las próximas mesas de final hasta que devuelvas el material prestado.

Saludos,
Biblioteca ISFD 210
                '''
            
            send_mail(
                subject='Sanción aplicada - Biblioteca ISFD 210',
                message=mensaje_email,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[sancion.usuario.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error enviando email de sanción: {e}")
        
        tipo_texto = "DOCENTE" if tipo_usuario_sancion == 'docente' else "ALUMNO"
        messages.success(
            request, 
            f"Sanción {tipo_texto} confirmada para {sancion.usuario.get_full_name()}. "
            f"Se ha enviado una notificación por email."
        )
        return redirect('gestionar_sanciones')
    
    return render(request, 'libros/confirmar_sancion.html', {'sancion': sancion})

# MODIFICAR la función cancelar_sancion existente para usar la nueva
@user_passes_test(es_bibliotecaria)
def cancelar_sancion(request, sancion_id):
    """
    Redirige a la nueva función con modal
    """
    return cancelar_sancion_con_modal(request, sancion_id)

# Modificar la función finalizar_prestamo para manejar sanciones
@user_passes_test(es_bibliotecaria)
def finalizar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id)
    
    if prestamo.estado not in ['aprobado', 'vencido']:
        messages.error(request, f"El préstamo no puede ser finalizado porque su estado actual es {prestamo.get_estado_display()}.")
        return redirect('gestionar_prestamos')
    
    prestamo.estado = 'finalizado'
    prestamo.fecha_devolucion_real = timezone.now()
    prestamo.save()
    
    # Devolver el libro al estado disponible
    libro = prestamo.libro
    libro.estado = 'Disponible'
    libro.save()
    
    # Cancelar automáticamente las sanciones relacionadas con este préstamo
    sanciones_activas = prestamo.sanciones.filter(estado__in=['pendiente', 'confirmada'])
    for sancion in sanciones_activas:
        sancion.estado = 'cumplida'
        sancion.fecha_finalizacion = timezone.now()
        sancion.observaciones_bibliotecaria = "Sanción finalizada automáticamente por devolución del material"
        sancion.save()
    
    if sanciones_activas.exists():
        messages.success(request, f"Préstamo finalizado y {sanciones_activas.count()} sanción(es) cancelada(s) automáticamente.")
    else:
        messages.success(request, f"El préstamo del libro '{prestamo.libro.titulo}' ha sido finalizado.")
    
    return redirect('gestionar_prestamos')

# Vista para que los usuarios vean sus sanciones
@login_required
def mis_sanciones(request):
    """Vista para que los usuarios vean sus sanciones"""
    sanciones = Sancion.objects.filter(usuario=request.user).order_by('-fecha_creacion')
    sanciones_activas = sanciones.filter(estado='confirmada')
    
    context = {
        'sanciones': sanciones,
        'sanciones_activas': sanciones_activas,
        'tiene_sanciones_activas': sanciones_activas.exists(),
    }
    return render(request, 'libros/mis_sanciones.html', context)


@user_passes_test(es_bibliotecaria)
def cancelar_sancion_con_modal(request, sancion_id):
    """
    Cancela una sanción con opción de devolver libro al catálogo
    """
    sancion = get_object_or_404(Sancion, id_sancion=sancion_id)
    
    if sancion.estado not in ['pendiente', 'confirmada']:
        messages.error(request, f"La sanción no puede ser cancelada porque su estado actual es {sancion.get_estado_display()}.")
        return redirect('gestionar_sanciones')
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '')
        devolver_libro = request.POST.get('devolver_libro') == 'on'
        
        # Cancelar la sanción
        sancion.estado = 'cancelada'
        sancion.observaciones_bibliotecaria = f"CANCELADA: {motivo_cancelacion}"
        sancion.fecha_finalizacion = timezone.now()
        
        if devolver_libro:
            # Devolver libro al catálogo
            sancion.libro_devuelto_catalogo = True
            sancion.fecha_devolucion_catalogo = timezone.now()
            
            # Cambiar estado del libro a disponible
            libro = sancion.prestamo.libro
            libro.estado = 'Disponible'
            libro.save()
            
            # También finalizar el préstamo si no está finalizado
            prestamo = sancion.prestamo
            if prestamo.estado in ['vencido', 'aprobado']:
                prestamo.estado = 'finalizado'
                prestamo.fecha_devolucion_real = timezone.now()
                prestamo.save()
            
            messages.success(request, f"Sanción cancelada y libro '{libro.titulo}' devuelto al catálogo.")
        else:
            sancion.libro_devuelto_catalogo = False
            messages.success(request, f"Sanción cancelada. El libro permanece fuera del catálogo.")
        
        sancion.save()
        return redirect('gestionar_sanciones')
    
    context = {
        'sancion': sancion,
        'libro': sancion.prestamo.libro,
        'usuario': sancion.usuario,
    }
    return render(request, 'libros/cancelar_sancion_con_modal.html', context)

@user_passes_test(es_bibliotecaria)
def devolver_libro_catalogo(request, sancion_id):
    """
    Devuelve un libro al catálogo desde una sanción cancelada
    """
    sancion = get_object_or_404(Sancion, id_sancion=sancion_id)
    
    if sancion.estado != 'cancelada':
        messages.error(request, "Solo se pueden devolver libros de sanciones canceladas.")
        return redirect('gestionar_sanciones')
    
    if sancion.libro_devuelto_catalogo:
        messages.warning(request, "Este libro ya fue devuelto al catálogo.")
        return redirect('gestionar_sanciones')
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        
        # Marcar libro como devuelto al catálogo
        sancion.libro_devuelto_catalogo = True
        sancion.fecha_devolucion_catalogo = timezone.now()
        
        # Agregar observaciones
        if observaciones:
            obs_anterior = sancion.observaciones_bibliotecaria or ""
            sancion.observaciones_bibliotecaria = f"{obs_anterior}\nLibro devuelto al catálogo: {observaciones}".strip()
        
        sancion.save()
        
        # Cambiar estado del libro a disponible
        libro = sancion.prestamo.libro
        libro.estado = 'Disponible'
        libro.save()
        
        # Finalizar el préstamo si no está finalizado
        prestamo = sancion.prestamo
        if prestamo.estado in ['vencido', 'aprobado']:
            prestamo.estado = 'finalizado'
            prestamo.fecha_devolucion_real = timezone.now()
            prestamo.save()
        
        messages.success(request, f"El libro '{libro.titulo}' ha sido devuelto al catálogo exitosamente.")
        return redirect('gestionar_sanciones')
    
    context = {
        'sancion': sancion,
        'libro': sancion.prestamo.libro,
        'usuario': sancion.usuario,
    }
    return render(request, 'libros/devolver_libro_catalogo.html', context)


@login_required
def verificar_sanciones_criticas(request):
    """Vista para mostrar cuando un docente tiene 3+ sanciones"""
    if not request.user.tiene_sanciones_criticas():
        return redirect('pantalla_principal')
    
    sanciones_docente = request.user.sanciones.filter(
        tipo_usuario_sancion='docente',
        estado='confirmada'
    ).order_by('-fecha_confirmacion')
    
    context = {
        'usuario': request.user,
        'sanciones_docente': sanciones_docente,
        'total_sanciones': sanciones_docente.count(),
    }
    return render(request, 'libros/sanciones_criticas.html', context)

@bibliotecaria_required
def resetear_acumulado_docente(request, usuario_id):
    """Resetea el acumulado de sanciones de un docente tras cumplir compensación"""
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        
        # Marcar las sanciones como cumplidas pero mantener historial
        sanciones_docente = usuario.sanciones.filter(
            tipo_usuario_sancion='docente',
            estado='confirmada'
        )
        
        for sancion in sanciones_docente:
            sancion.acumulado_sanciones = 0  # Resetear contador
            sancion.observaciones_bibliotecaria += f"\n[COMPENSACIÓN CUMPLIDA] {observaciones}"
            sancion.save()
        
        messages.success(
            request, 
            f"Acumulado de sanciones reseteado para {usuario.get_full_name()}. "
            f"Puede volver a solicitar préstamos normalmente."
        )
        return redirect('gestionar_sanciones')
    
    context = {
        'usuario': usuario,
        'sanciones_count': usuario.sanciones.filter(
            tipo_usuario_sancion='docente', 
            estado='confirmada'
        ).count()
    }
    return render(request, 'libros/resetear_acumulado.html', context)

@login_required
def mis_prestamos_activos(request):
    """Vista para mostrar los préstamos activos del usuario"""
    prestamos_activos = request.user.get_prestamos_activos().order_by('-fecha_solicitud')
    
    context = {
        'prestamos_activos': prestamos_activos,
        'total_prestamos': prestamos_activos.count(),
        'prestamos_disponibles': request.user.get_prestamos_disponibles()
    }
    
    return render(request, 'libros/mis_prestamos_activos.html', context)

@login_required
def extender_prestamo(request, prestamo_id):
    """Vista para que los docentes extiendan sus préstamos"""
    prestamo = get_object_or_404(Prestamo, id_prestamo=prestamo_id, usuario=request.user)
    
    # Verificar que sea docente y pueda extender
    if request.user.perfil != 'docente':
        messages.error(request, "Solo los docentes pueden extender préstamos.")
        return redirect('prestamos_solicitados')
    
    if not prestamo.puede_extender_prestamo():
        messages.error(request, "Este préstamo no puede ser extendido.")
        return redirect('prestamos_solicitados')
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', 'Extensión solicitada por el docente')
        
        if prestamo.extender_prestamo(motivo):
            # Enviar email de confirmación
            try:
                send_mail(
                    subject='Préstamo Extendido - Biblioteca ISFD 210',
                    message=f'''
Estimado/a {request.user.get_full_name()},

Tu préstamo ha sido extendido exitosamente:

📚 Libro: {prestamo.libro.titulo}
📅 Nueva fecha de devolución: {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y a las %H:%M')}
📝 Motivo: {motivo}

La extensión es de 15 días adicionales. Recuerda devolver el material en la nueva fecha límite.

Saludos,
Biblioteca ISFD 210
                    ''',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[request.user.email],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Error enviando email de extensión: {e}")
            
            messages.success(request, f"Préstamo extendido hasta el {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y a las %H:%M')}.")
        else:
            messages.error(request, "No se pudo extender el préstamo.")
        
        return redirect('prestamos_solicitados')
    
    context = {
        'prestamo': prestamo,
        'nueva_fecha': prestamo.fecha_devolucion_programada + datetime.timedelta(days=15)
    }
    
    return render(request, 'libros/extender_prestamo.html', context)

# Función para enviar avisos de vencimiento (mejorada)
def enviar_avisos_vencimiento():
    """
    Envía emails de aviso 24 horas antes del vencimiento
    """
    ahora = timezone.now()
    en_24_horas = ahora + datetime.timedelta(hours=24)
    
    # Préstamos que vencen en las próximas 24 horas
    prestamos_por_vencer = Prestamo.objects.filter(
        estado='aprobado',
        fecha_devolucion_programada__range=(ahora, en_24_horas)
    )
    
    for prestamo in prestamos_por_vencer:
        try:
            # Mensaje diferente para docentes (que pueden extender)
            if prestamo.usuario and prestamo.usuario.perfil == 'docente' and prestamo.puede_extender_prestamo():
                mensaje = f'''
Estimado/a {prestamo.nombre_usuario},

Tu préstamo vence en menos de 24 horas:

📚 "{prestamo.libro.titulo}"
✍️ Autor: {prestamo.libro.autor}
⏰ Vence: {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y a las %H:%M')}

Como eres docente, puedes EXTENDER este préstamo por 15 días más desde tu panel de préstamos en el sistema.

🔗 Accede al sistema: http://127.0.0.1:8000/libros/prestamos-solicitados/

Si no necesitas extenderlo, por favor acércate a la biblioteca para realizar la devolución.

Saludos,
Biblioteca ISFD 210
                '''
            else:
                mensaje = f'''
Estimado/a {prestamo.nombre_usuario},

Te recordamos que en menos de 24 horas vence el plazo para devolver el libro:

📚 "{prestamo.libro.titulo}"
✍️ Autor: {prestamo.libro.autor}
⏰ Vence: {prestamo.fecha_devolucion_programada.strftime('%d/%m/%Y a las %H:%M')}

Por favor, acércate a la biblioteca para realizar la devolución y evitar sanciones.

Saludos,
Biblioteca ISFD 210
                '''
            
            send_mail(
                subject='⚠️ Préstamo por Vencer - Biblioteca ISFD 210',
                message=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[prestamo.email_usuario],
                fail_silently=False,
            )
            
            print(f"[EMAIL] Aviso de vencimiento enviado a {prestamo.email_usuario}")
            
        except Exception as e:
            print(f"Error enviando email a {prestamo.email_usuario}: {e}")

# Tarea programada para verificar vencimientos (agregar al final del archivo)
def verificar_y_notificar_vencimientos():
    """
    Función para ejecutar diariamente - verifica vencimientos y envía emails
    """
    print(f"[CRON] Verificando vencimientos a las {timezone.now()}")
    
    # Enviar avisos de vencimiento
    enviar_avisos_vencimiento()
    
    # Verificar préstamos vencidos
    verificar_prestamos_vencidos()



@user_passes_test(es_bibliotecaria)
def exportar_bajas_excel(request):
    """Vista para exportar libros dados de baja a Excel"""
    try:
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro de Bajas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'ID Libro', 'Título', 'Autor', 'Editorial', 'Clasificación CDU', 
            'Número Inventario', 'Sede', 'Motivo de Baja', 'Descripción', 
            'Número Ejemplar', 'Fecha de Baja'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos de libros dados de baja
        libros_baja = Libro.objects.filter(estado='No disponible').order_by('-id_libro')
        
        for row, libro in enumerate(libros_baja, 2):
            ws.cell(row=row, column=1, value=libro.id_libro)
            ws.cell(row=row, column=2, value=libro.titulo)
            ws.cell(row=row, column=3, value=libro.autor)
            ws.cell(row=row, column=4, value=libro.editorial)
            ws.cell(row=row, column=5, value=libro.clasificacion_cdu)
            ws.cell(row=row, column=6, value=libro.num_inventario)
            ws.cell(row=row, column=7, value=libro.sede)
            ws.cell(row=row, column=8, value=libro.motivo_baja or 'No especificado')
            ws.cell(row=row, column=9, value=libro.descripcion or 'Sin descripción')
            ws.cell(row=row, column=10, value=libro.num_ejemplar)
            # Nota: Si no tienes fecha de baja, puedes agregar un campo al modelo o usar una fecha por defecto
            ws.cell(row=row, column=11, value=timezone.now().strftime('%d/%m/%Y'))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=registro_bajas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar registro de bajas: {str(e)}')
        return redirect('registro_de_bajas')
    
# Vistas para migraciones de localhost (sqlite) a Postgres estan comentadas porque una vez que se hace no se tiene que volver a hacer. 
# Dejar la url abierta puede ser un peligro porque acceden y te crean superusuario. Se comentan los metodos.


# @csrf_exempt
# def crear_admin_temporal(request):
#     """Vista temporal para crear admin - ELIMINAR DESPUÉS"""
#     from django.db import connection
    
#     # Verificar si hay tablas
#     try:
#         with connection.cursor() as cursor:
#             cursor.execute("SELECT to_regclass('public.libros_usuario')")
#             tabla_existe = cursor.fetchone()[0] is not None
#     except:
#         return HttpResponse("Error: No se pudo verificar las tablas. Ejecuta primero /libros/ejecutar-migraciones/")
    
#     if not tabla_existe:
#         return HttpResponse("Error: Las tablas no existen. Ve a <a href='/libros/ejecutar-migraciones/'>/libros/ejecutar-migraciones/</a>")
    
#     try:
#         if Usuario.objects.filter(perfil='bibliotecaria').exists():
#             return HttpResponse("Ya existe una bibliotecaria.")
#     except:
#         pass
    
#     if request.method == 'POST':
#         try:
#             usuario = Usuario.objects.create_superuser(
#                 dni=request.POST['dni'],
#                 password=request.POST['password'],
#                 nombre=request.POST['nombre'],
#                 apellido=request.POST['apellido'],
#                 email=request.POST['email']
#             )
#             return HttpResponse(f"✅ Admin creado: {usuario.dni}. ELIMINA estas vistas del código ahora.")
#         except Exception as e:
#             return HttpResponse(f"❌ Error: {e}")
    
#     return HttpResponse('''
#         <form method="post">
#             DNI: <input name="dni" required><br><br>
#             Password: <input type="password" name="password" required><br><br>
#             Nombre: <input name="nombre" required><br><br>
#             Apellido: <input name="apellido" required><br><br>
#             Email: <input name="email" type="email" required><br><br>
#             <button>Crear Admin</button>
#         </form>
#     ''')


# def ejecutar_migraciones(request):
#     """Vista temporal para ejecutar migraciones - ELIMINAR DESPUÉS"""
#     output = StringIO()
#     try:
#         # Ejecutar migraciones
#         call_command('migrate', stdout=output)
#         resultado = output.getvalue()
        
#         return HttpResponse(f"<pre>Migraciones ejecutadas:\n\n{resultado}</pre>")
#     except Exception as e:
#         return HttpResponse(f"<pre>Error ejecutando migraciones:\n\n{str(e)}</pre>")